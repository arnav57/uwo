import openpyxl, os, sys, re

## HELPER FUNCTIONS

LEC_FOLDER_NAME 	= 'lec'
PROJ_FOLDER_NAME 	= 'proj'
TEST_FOLDER_NAME	= 'test'
LAB_FOLDER_NAME		= 'lab'
TUT_FOLDER_NAME		= 'tut'

LEC_HANDWRITTEN_FOLDER_NAME = 'pen'
LEC_MD_FOLDER_NAME			= 'md'

BASE_MAKEFILE = '''
# PANDOC ENGINE COMMANDS
PDF_ENGINE := xelatex
MD_DIR     := ./md/
MD_FILES   := $(wildcard $(MD_DIR)/*.md)

# COURSE NAMES
MAKEFILE_DIR  := $(dir $(abspath $(lastword $(MAKEFILE_LIST))))
OUTER_PATH    := $(dir $(patsubst %/,%,$(MAKEFILE_DIR)))
COURSE_NAME   := $(notdir $(patsubst %/,%,$(OUTER_PATH)))
PDF_OUT_NAME  := $(COURSE_NAME).pdf
PDF_CONFIG    := $(MAKEFILE_DIR)/pdf_config.yaml


info:
	@echo "Output Names  : $(OUTER_NAME).pdf"
	@echo "$(MD_FILES)"

pdf: $(MD_FILES)
	pandoc $(MD_FILES) \\
	--from markdown \\
	--pdf-engine $(PDF_ENGINE) \\
	--metadata-file=$(PDF_CONFIG) \\
	--to pdf \\
	-o $(PDF_OUT_NAME)
'''

def assemble_course_folder_name(course_code):
	parts = course_code.split()
	# parts[0]  -> course dept (ECE, AISE, etc)
	# parts[1]  -> course num
	# parts[2:] -> discarded

	course_name = f"{parts[0].lower()}_{parts[1]}"
	return course_name

def check_and_create_course_folder(search_path, course_folder_name):
	total_path = os.path.join(search_path, course_folder_name)

	if not os.path.isdir(total_path):
		os.makedirs(total_path)
		print(f"Created {total_path}")
		return True
	else:
		print(f"Path {total_path} already exists")
		return False

def create_course_structure(search_path, course_folder_name, has_lab, has_tut):
	course_path = os.path.join(search_path, course_folder_name)

	# each course should have the following folders by default
	# -> lec (lecture notes)
	# -> proj (projects, assignments)
	# -> test (quizzes, tests, etc.)

	# create folders
	os.makedirs( os.path.join(course_path, LEC_FOLDER_NAME ) )
	os.makedirs( os.path.join(course_path, PROJ_FOLDER_NAME) )
	os.makedirs( os.path.join(course_path, TEST_FOLDER_NAME) )

	# create inner llecture-folders
	os.makedirs( os.path.join(course_path, LEC_FOLDER_NAME, LEC_HANDWRITTEN_FOLDER_NAME ) )
	os.makedirs( os.path.join(course_path, LEC_FOLDER_NAME, LEC_MD_FOLDER_NAME 			) )

	if (has_lab.upper().rstrip() == "Y"):
		os.makedirs( os.path.join(course_path, LAB_FOLDER_NAME ) )

	if (has_tut.upper().rstrip() == "Y"):
		os.makedirs( os.path.join(course_path, TUT_FOLDER_NAME ) )

def create_base_readme(search_path, course_folder_name, course_name, course_desc, has_lab, has_tut):
	course_path = os.path.join(search_path, course_folder_name)

	base_readme_text = f'''
# {course_name.upper()}
{course_desc}
----
- {LEC_FOLDER_NAME}: handwritten, markdown, and pdf versions of lecture notes
- {PROJ_FOLDER_NAME}: projects and assignments
- {TEST_FOLDER_NAME}: quizzes, tests, and assessment type stuff'''
	extra_readme_text = ""

	if (has_lab.upper().rstrip() == "Y"):
		extra_readme_text = extra_readme_text + f"\n- {LAB_FOLDER_NAME}: lab stuff"

	if (has_tut.upper().rstrip() == "Y"):
		extra_readme_text = extra_readme_text + f"\n- {TUT_FOLDER_NAME}: tutorial stuff"
	total_readme_text = base_readme_text + extra_readme_text

	## create the readme.md file and print the text
	file = open(os.path.join(course_path, "README.md"), 'w')
	print(total_readme_text, file=file)

def create_base_makefile(search_path, course_folder_name):
	course_path = os.path.join(search_path, course_folder_name)
	base_makefile_text = BASE_MAKEFILE

	## create the Makefile and print the text
	file = open(os.path.join(course_path, LEC_FOLDER_NAME, "Makefile"), 'w')
	print(base_makefile_text, file=file)

def create_base_yaml(search_path, course_folder_name, course_name, course_desc):
	YAML_TEXT = rf'''
title: "{course_name}"
subtitle: "{course_desc}"
author: "Arnav Goyal"
date: "2025-2026"
toc: true
numbersections: true
fontsize: 12pt
mainfont: "TeX Gyre Pagella"
monofont: "JetBrains Mono"
header-includes:
  - \input{{../../../assets/preamble.tex}}
  - \usepackage{{etoolbox}}
  - \preto{{\tableofcontents}}{{\clearpage}}
  - \apptocmd{{\tableofcontents}}{{\clearpage}}{{}}{{}}
'''

	## create the YAML file and print the text
	course_path = os.path.join(search_path, course_folder_name)
	file = open(os.path.join(course_path, LEC_FOLDER_NAME, "pdf_config.yaml"), 'w')
	print(YAML_TEXT, file=file)

## CONSTANTS
COURSE_DIR 			= os.path.abspath("../main/")
COURSE_LIST_XLSX 	= os.path.abspath("../course_list.xlsx")
COL_HEADER_ROW_NUM  = 2

COURSE_CODE_COL 	= 0
COURSE_DESC_COL 	= 1
COURSE_HAS_LAB_COL 	= 2
COURSE_HAS_TUT_COL 	= 3


## READ IN COURSES EXCEL SHEET
wb = openpyxl.load_workbook(COURSE_LIST_XLSX)
sheet = wb['Year 4']

new_course_folder_names = []
for row in sheet.iter_rows(min_row=COL_HEADER_ROW_NUM, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
	row_num = row[0].row

	current_course_code 	= row[COURSE_CODE_COL].value
	current_course_desc 	= row[COURSE_DESC_COL].value
	current_course_has_lab 	= row[COURSE_HAS_LAB_COL].value
	current_course_has_tut 	= row[COURSE_HAS_TUT_COL].value

	# get properly formatted course code names
	current_folder_name = assemble_course_folder_name(current_course_code)

	# if folder doesnt exist, its new so add it to the new courses list
	if check_and_create_course_folder(COURSE_DIR, current_folder_name):
		create_course_structure(COURSE_DIR, current_folder_name, current_course_has_lab, current_course_has_tut)
		create_base_readme(COURSE_DIR, current_folder_name, current_course_code, current_course_desc, current_course_has_lab, current_course_has_tut)
		create_base_makefile(COURSE_DIR, current_folder_name)
		create_base_yaml(COURSE_DIR, current_folder_name, current_course_code, current_course_desc)



